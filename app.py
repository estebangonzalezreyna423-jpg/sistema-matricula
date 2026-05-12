from flask import Flask, render_template, request, redirect, session
import sqlite3
from datetime import datetime
import pandas as pd
from flask import send_file

app = Flask(__name__)
app.secret_key = "clave_secreta"


def get_db():
    return sqlite3.connect("colegio.db")


def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT,
        rol TEXT DEFAULT 'secretaria'
    )
    """)

    try:
        cur.execute("ALTER TABLE usuarios ADD COLUMN rol TEXT DEFAULT 'secretaria'")
    except:
        pass

    cur.execute("""
    INSERT OR IGNORE INTO usuarios (id, username, password, rol)
    VALUES (1, 'admin', '1234', 'admin')
    """)

    cur.execute("""
        UPDATE usuarios
        SET rol='admin'
        WHERE username='admin'
    """)

    cur.execute("""
        INSERT OR IGNORE INTO usuarios (username, password, rol)
        VALUES ('secretaria', '1234', 'secretaria')
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS alumnos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombres TEXT NOT NULL,
        apellidos TEXT NOT NULL,
        dni TEXT,
        grado TEXT NOT NULL,
        seccion TEXT NOT NULL,
        apoderado TEXT,
        telefono TEXT,
        estado TEXT DEFAULT 'Activo'
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS matriculas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alumno_id INTEGER NOT NULL,
        fecha TEXT NOT NULL,
        estado TEXT DEFAULT 'Matriculado',
        FOREIGN KEY (alumno_id) REFERENCES alumnos(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS pagos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alumno_id INTEGER NOT NULL,
        concepto TEXT NOT NULL,
        monto REAL DEFAULT 0,
        estado TEXT DEFAULT 'Pendiente',
        fecha_pago TEXT,
        fecha_vencimiento TEXT,
        FOREIGN KEY (alumno_id) REFERENCES alumnos(id)
    )
    """)

    try:
        cur.execute("ALTER TABLE matriculas ADD COLUMN monto_matricula REAL DEFAULT 0")
    except:
        pass

    try:
        cur.execute("ALTER TABLE matriculas ADD COLUMN monto_pension REAL DEFAULT 0")
    except:
        pass

    try:
        cur.execute("ALTER TABLE matriculas ADD COLUMN descuento TEXT")
    except:
        pass

    try:
        cur.execute("ALTER TABLE usuarios ADD COLUMN rol TEXT DEFAULT 'secretaria'")
    except:
        pass

    try:
        cur.execute("ALTER TABLE pagos ADD COLUMN fecha_vencimiento TEXT")
    except:
        pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS boletas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_boleta TEXT,
        alumno_id INTEGER,
        pago_id INTEGER,
        concepto TEXT,
        monto_recibido REAL,
        metodo_pago TEXT,
        fecha_pago_real TEXT,
        fecha_confirmacion TEXT,
        fecha_registro TEXT,
        observacion TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS boleta_detalles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        boleta_id INTEGER,
        pago_id INTEGER,
        monto_aplicado REAL
    )
    """)
    try:
       cur.execute("ALTER TABLE boletas ADD COLUMN numero_boleta TEXT")
    except:
        pass

    conn.commit()
    conn.close()


init_db()


@app.context_processor
def injectar_usuario():
    return {
        "usuario_actual": session.get("user"),
        "rol_actual": session.get("rol")
    }


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form["username"]
        password = request.form["password"]

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, username, password, rol
            FROM usuarios
            WHERE username=? AND password=?
        """, (user, password))
        usuario = cur.fetchone()
        conn.close()

        if usuario:
            session["user"] = usuario[1]
            session["rol"] = usuario[3]
            return redirect("/dashboard")

    return render_template("login.html")


@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM alumnos")
    total_alumnos = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM alumnos WHERE estado='Activo'")
    alumnos_activos = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM alumnos WHERE estado='Retirado'")
    alumnos_retirados = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM matriculas")
    total_matriculados = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM pagos WHERE estado='Pendiente'")
    pagos_pendientes = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM pagos WHERE estado='Pagado'")
    pagos_pagados = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM pagos WHERE estado='Exonerado'")
    pagos_exonerados = cur.fetchone()[0]

    try:
        cur.execute("SELECT SUM(monto_recibido) FROM boletas")
        total_recaudado = cur.fetchone()[0] or 0
    except:
        total_recaudado = 0

    cur.execute("SELECT SUM(monto) FROM pagos WHERE estado='Pendiente'")
    total_pendiente = cur.fetchone()[0] or 0

    try:
        cur.execute("SELECT COUNT(*) FROM boletas")
        total_boletas = cur.fetchone()[0]
    except:
        total_boletas = 0

    cur.execute("""
        SELECT COUNT(DISTINCT alumno_id)
        FROM pagos
        WHERE estado='Pendiente'
    """)
    alumnos_deudores = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*)
        FROM alumnos
        WHERE id IN (SELECT alumno_id FROM matriculas)
        AND id NOT IN (
            SELECT alumno_id FROM pagos WHERE estado='Pendiente'
        )
    """)
    alumnos_al_dia = cur.fetchone()[0]

    conn.close()

    return render_template(
        "dashboard.html",
        total_alumnos=total_alumnos,
        alumnos_activos=alumnos_activos,
        alumnos_retirados=alumnos_retirados,
        total_matriculados=total_matriculados,
        pagos_pendientes=pagos_pendientes,
        pagos_pagados=pagos_pagados,
        pagos_exonerados=pagos_exonerados,
        total_recaudado=total_recaudado,
        total_pendiente=total_pendiente,
        total_boletas=total_boletas,
        alumnos_deudores=alumnos_deudores,
        alumnos_al_dia=alumnos_al_dia
    )


@app.route("/alumnos", methods=["GET", "POST"])
def alumnos():
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    if request.method == "POST":
        nombres = request.form["nombres"]
        apellidos = request.form["apellidos"]
        dni = request.form["dni"]
        grado = request.form["grado"]
        seccion = request.form["seccion"]
        apoderado = request.form["apoderado"]
        telefono = request.form["telefono"]

        cur.execute("""
            INSERT INTO alumnos 
            (nombres, apellidos, dni, grado, seccion, apoderado, telefono)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (nombres, apellidos, dni, grado, seccion, apoderado, telefono))

        conn.commit()
        conn.close()
        return redirect("/alumnos")

    buscar = request.args.get("buscar", "")
    grado = request.args.get("grado", "")
    seccion = request.args.get("seccion", "")

    query = "SELECT * FROM alumnos WHERE 1=1"
    params = []

    if buscar:
        query += " AND (nombres LIKE ? OR apellidos LIKE ? OR dni LIKE ?)"
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if grado:
        query += " AND grado=?"
        params.append(grado)

    if seccion:
        query += " AND seccion=?"
        params.append(seccion)

    query += " ORDER BY grado, seccion, apellidos"

    cur.execute(query, params)
    alumnos_lista = cur.fetchall()
    conn.close()

    return render_template("alumnos.html", alumnos=alumnos_lista)


@app.route("/matricula")
def matricula():
    if "user" not in session:
        return redirect("/")

    buscar = request.args.get("buscar", "")
    grado = request.args.get("grado", "")
    seccion = request.args.get("seccion", "")

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT 
            alumnos.id,
            alumnos.nombres,
            alumnos.apellidos,
            alumnos.grado,
            alumnos.seccion,
            CASE 
                WHEN matriculas.id IS NOT NULL THEN 'Sí'
                ELSE 'No'
            END AS matriculado
        FROM alumnos
        LEFT JOIN matriculas ON alumnos.id = matriculas.alumno_id
        WHERE 1=1
    """

    params = []

    if buscar:
        query += " AND (alumnos.nombres LIKE ? OR alumnos.apellidos LIKE ? OR alumnos.dni LIKE ?)"
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if grado:
        query += " AND alumnos.grado=?"
        params.append(grado)

    if seccion:
        query += " AND alumnos.seccion=?"
        params.append(seccion)

    query += " ORDER BY alumnos.grado, alumnos.seccion, alumnos.apellidos"

    cur.execute(query, params)
    alumnos_lista = cur.fetchall()
    conn.close()

    return render_template("matricula.html", alumnos=alumnos_lista)


@app.route("/matricular/<int:alumno_id>", methods=["GET", "POST"])
def matricular_alumno(alumno_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM alumnos WHERE id=?", (alumno_id,))
    alumno = cur.fetchone()

    cur.execute("SELECT id FROM matriculas WHERE alumno_id=?", (alumno_id,))
    ya_matriculado = cur.fetchone()

    if ya_matriculado:
        conn.close()
        return redirect("/pagos/" + str(alumno_id))

    if request.method == "POST":

        monto_matricula = float(request.form["monto_matricula"])
        monto_pension = float(request.form["monto_pension"])
        descuento = request.form["descuento"]

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cur.execute("""
            INSERT INTO matriculas
            (alumno_id, fecha, monto_matricula, monto_pension, descuento)
            VALUES (?, ?, ?, ?, ?)
        """, (
            alumno_id,
            fecha,
            monto_matricula,
            monto_pension,
            descuento
        ))

        pagos = [
            ("Matrícula", monto_matricula, fecha[:10]),
            ("Agenda", 20, fecha[:10]),
            ("Cuota de ingreso", 100, fecha[:10]),

            ("Marzo", monto_pension, "2026-03-31"),
            ("Abril", monto_pension, "2026-04-30"),
            ("Mayo", monto_pension, "2026-05-31"),
            ("Junio", monto_pension, "2026-06-30"),
            ("Julio", monto_pension, "2026-07-31"),
            ("Agosto", monto_pension, "2026-08-31"),
            ("Septiembre", monto_pension, "2026-09-30"),
            ("Octubre", monto_pension, "2026-10-31"),
            ("Noviembre", monto_pension, "2026-11-30"),
            ("Diciembre", monto_pension, "2026-12-31")
        ]

        for concepto, monto, fecha_vencimiento in pagos:

            cur.execute("""
                SELECT id
                FROM pagos
                WHERE alumno_id=? AND concepto=?
            """, (alumno_id, concepto))

            existe = cur.fetchone()

            if not existe:

                cur.execute("""
                    INSERT INTO pagos
                    (alumno_id, concepto, monto, estado, fecha_vencimiento)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    alumno_id,
                    concepto,
                    monto,
                    "Pendiente",
                    fecha_vencimiento
                ))

        conn.commit()
        conn.close()

        return redirect("/pagos/" + str(alumno_id))

    conn.close()

    return render_template(
        "confirmar_matricula.html",
        alumno=alumno
    )


@app.route("/pagos/<int:alumno_id>")
def ver_pagos(alumno_id):

    if "user" not in session:
        return redirect("/")

    buscar = request.args.get("buscar", "")
    estado = request.args.get("estado", "")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombres, apellidos
        FROM alumnos
        WHERE id=?
    """, (alumno_id,))

    alumno = cur.fetchone()

    query = """
        SELECT 
            pagos.id,
            pagos.concepto,
            pagos.monto,
            pagos.estado,
            pagos.fecha_pago,

            (
                SELECT boleta_detalles.boleta_id
                FROM boleta_detalles
                WHERE boleta_detalles.pago_id = pagos.id
                ORDER BY boleta_detalles.id DESC
                LIMIT 1
            ) AS boleta_id

        FROM pagos
        WHERE pagos.alumno_id=?
    """

    params = [alumno_id]

    if buscar:
        query += " AND pagos.concepto LIKE ?"
        params.append(f"%{buscar}%")

    if estado:
        query += " AND pagos.estado=?"
        params.append(estado)

    query += " ORDER BY pagos.id ASC"

    cur.execute(query, params)
    pagos = cur.fetchall()

    cur.execute("""
        SELECT SUM(monto)
        FROM pagos
        WHERE alumno_id=?
    """, (alumno_id,))
    pendiente = cur.fetchone()[0] or 0

    try:
        cur.execute("""
            SELECT SUM(monto_recibido)
            FROM boletas
            WHERE alumno_id=?
        """, (alumno_id,))
        pagado = cur.fetchone()[0] or 0
    except:
        pagado = 0

    total = pendiente + pagado

    conn.close()

    return render_template(
        "pagos.html",
        pagos=pagos,
        alumno=alumno,
        total=total,
        pagado=pagado,
        pendiente=pendiente,
        buscar=buscar,
        estado=estado
    )


@app.route("/pagar/<int:pago_id>", methods=["GET", "POST"])
def pagar(pago_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS boletas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_boleta TEXT,
        alumno_id INTEGER,
        pago_id INTEGER,
        concepto TEXT,
        monto_recibido REAL,
        metodo_pago TEXT,
        fecha_pago_real TEXT,
        fecha_confirmacion TEXT,
        fecha_registro TEXT,
        observacion TEXT
    )
    """)

    cur.execute("""
        SELECT pagos.id, pagos.alumno_id, pagos.concepto, pagos.monto, pagos.estado,
               alumnos.nombres, alumnos.apellidos
        FROM pagos
        JOIN alumnos ON pagos.alumno_id = alumnos.id
        WHERE pagos.id=?
    """, (pago_id,))

    pago = cur.fetchone()

    if not pago:
        conn.close()
        return redirect("/matricula")

    if request.method == "POST":

        monto_recibido = float(request.form["monto_recibido"])
        numero_boleta = request.form["numero_boleta"]
        metodo_pago = request.form["metodo_pago"]
        fecha_pago_real = request.form["fecha_pago_real"]
        fecha_confirmacion = request.form["fecha_confirmacion"]
        observacion = request.form["observacion"]
        descripcion_boleta = request.form.get("descripcion_boleta", "")

        alumno_id = pago[1]
        saldo = monto_recibido
        fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # CREAR BOLETA
        cur.execute("""
            INSERT INTO boletas
            (
                numero_boleta,
                alumno_id,
                pago_id,
                concepto,
                monto_recibido,
                metodo_pago,
                fecha_pago_real,
                fecha_confirmacion,
                fecha_registro,
                observacion
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            numero_boleta,
            alumno_id,
            pago_id,
            pago[2],
            monto_recibido,
            metodo_pago,
            fecha_pago_real,
            fecha_confirmacion,
            fecha_registro,
            observacion + "" + descripcion_boleta
        ))

        boleta_id = cur.lastrowid

        # BUSCAR PAGOS PENDIENTES
        cur.execute("""
            SELECT id, monto
            FROM pagos
            WHERE alumno_id=? 
            AND estado='Pendiente'
            AND id>=?
            ORDER BY id ASC
        """, (alumno_id, pago_id))

        pagos_pendientes = cur.fetchall()

        for p in pagos_pendientes:

            if saldo <= 0:
                break

            id_pago_actual = p[0]
            deuda_actual = p[1]

            # SI EL SALDO ALCANZA
            if saldo >= deuda_actual:

                monto_aplicado = deuda_actual
                saldo -= deuda_actual

                cur.execute("""
                    UPDATE pagos
                    SET monto=0,
                        estado='Pagado',
                        fecha_pago=?
                    WHERE id=?
                """, (fecha_confirmacion, id_pago_actual))

            # SI EL SALDO NO ALCANZA
            else:

                monto_aplicado = saldo
                nuevo_monto = deuda_actual - saldo
                saldo = 0

                cur.execute("""
                    UPDATE pagos
                    SET monto=?
                    WHERE id=?
                """, (nuevo_monto, id_pago_actual))

            # DETALLE DE BOLETA
            cur.execute("""
                INSERT INTO boleta_detalles
                (boleta_id, pago_id, monto_aplicado)
                VALUES (?, ?, ?)
            """, (
                boleta_id,
                id_pago_actual,
                monto_aplicado
            ))

        conn.commit()
        conn.close()

        return redirect("/pagos/" + str(alumno_id))

    conn.close()

    return render_template(
        "registrar_pago.html",
        pago=pago
    )


@app.route("/boleta/<int:boleta_id>")
def boleta(boleta_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            boletas.id,
            boletas.numero_boleta,
            boletas.concepto,
            boletas.monto_recibido,
            boletas.metodo_pago,
            boletas.fecha_pago_real,
            boletas.fecha_confirmacion,
            boletas.fecha_registro,
            boletas.observacion,
            alumnos.nombres,
            alumnos.apellidos,
            alumnos.grado,
            alumnos.seccion,
            alumnos.id,
            boletas.pago_id
        FROM boletas
        JOIN alumnos ON boletas.alumno_id = alumnos.id
        WHERE boletas.id=?
    """, (boleta_id,))

    boleta = cur.fetchone()

    if not boleta:
        conn.close()
        return redirect("/boletas")

    alumno_id = boleta[13]
    pago_id = boleta[14]

    cur.execute("""
        SELECT 
            boletas.id,
            boleta_detalles.monto_aplicado,
            boletas.metodo_pago,
            boletas.fecha_pago_real,
            boletas.fecha_confirmacion,
 
            SUBSTR(
                boletas.observacion,
                INSTR(boletas.observacion, '||') + 2
            )

        FROM boleta_detalles
        JOIN boletas ON boleta_detalles.boleta_id = boletas.id
        WHERE boleta_detalles.pago_id=?
        ORDER BY boletas.id ASC
    """, (pago_id,))

    historial = cur.fetchall()

    cur.execute("""
        SELECT concepto, monto, estado
        FROM pagos
        WHERE id=?
    """, (pago_id,))

    pago_actual = cur.fetchone()

    total_pagado = sum([h[1] for h in historial])
    pendiente = pago_actual[1]
    total_concepto = total_pagado + pendiente

    conn.close()

    return render_template(
        "boleta.html",
        boleta=boleta,
        historial=historial,
        pago_actual=pago_actual,
        total_pagado=total_pagado,
        pendiente=pendiente,
        total_concepto=total_concepto,
        alumno_id=alumno_id
    )

@app.route("/boletas_alumno/<int:alumno_id>")
def boletas_alumno(alumno_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombres, apellidos, grado, seccion
        FROM alumnos
        WHERE id=?
    """, (alumno_id,))
    alumno = cur.fetchone()

    cur.execute("""
        SELECT 
            id,
            numero_boleta,
            concepto,
            monto_recibido,
            metodo_pago,
            fecha_pago_real,
            fecha_confirmacion
        FROM boletas
        WHERE alumno_id=?
        ORDER BY id DESC
    """, (alumno_id,))
    boletas = cur.fetchall()

    conn.close()

    return render_template(
        "boletas_alumno.html",
        alumno=alumno,
        boletas=boletas
    )


@app.route("/boletas")
def boletas_general():

    if "user" not in session:
        return redirect("/")

    buscar = request.args.get("buscar", "")

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT DISTINCT 
            alumnos.id,
            alumnos.nombres,
            alumnos.apellidos,
            alumnos.grado,
            alumnos.seccion
        FROM alumnos
        JOIN boletas 
            ON alumnos.id = boletas.alumno_id
        WHERE 1=1
    """

    params = []

    if buscar:
        query += """
            AND (
                alumnos.nombres LIKE ?
                OR alumnos.apellidos LIKE ?
                OR boletas.numero_boleta LIKE ?
            )
        """

        params.extend([
            f"%{buscar}%",
            f"%{buscar}%",
            f"%{buscar}%"
        ])

    query += """
        ORDER BY 
            alumnos.grado,
            alumnos.seccion,
            alumnos.apellidos
    """

    cur.execute(query, params)
    alumnos = cur.fetchall()

    conn.close()

    return render_template(
        "boletas_general.html",
        alumnos=alumnos,
        buscar=buscar
    )

@app.route("/descargar_boletas_excel")
def descargar_boletas_excel():

    if "user" not in session:
        return redirect("/")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            boletas.numero_boleta,
            alumnos.nombres,
            alumnos.apellidos,
            alumnos.grado,
            alumnos.seccion,
            boletas.concepto,
            boletas.monto_recibido,
            boletas.metodo_pago,
            boletas.fecha_pago_real,
            boletas.fecha_confirmacion
        FROM boletas
        JOIN alumnos
            ON boletas.alumno_id = alumnos.id
        ORDER BY boletas.id DESC
    """)

    datos = cur.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Boletas"

    encabezados = [
        "N° BOLETA",
        "ALUMNO",
        "GRADO",
        "SECCIÓN",
        "CONCEPTO",
        "MONTO",
        "MÉTODO",
        "FECHA PAGO",
        "FECHA CONFIRMACIÓN"
    ]

    ws.append(encabezados)

    for d in datos:

        alumno = f"{d[1]} {d[2]}"

        fila = [
            d[0],
            alumno,
            d[3],
            d[4],
            d[5],
            d[6],
            d[7],
            d[8],
            d[9]
        ]

        ws.append(fila)

    azul = PatternFill("solid", fgColor="1E3A8A")
    blanco = Font(color="FFFFFF", bold=True)

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = azul
        cell.font = blanco
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:

        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    nombre_archivo = "boletas.xlsx"

    wb.save(nombre_archivo)

    return send_file(nombre_archivo, as_attachment=True)

@app.route("/deudores")
def deudores():
    if "user" not in session:
        return redirect("/")

    buscar = request.args.get("buscar", "")
    grado = request.args.get("grado", "")
    seccion = request.args.get("seccion", "")
    mes = request.args.get("mes", "")

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT 
            alumnos.id,
            alumnos.nombres,
            alumnos.apellidos,
            alumnos.grado,
            alumnos.seccion,
            COUNT(pagos.id) AS cantidad_pendiente,
            SUM(pagos.monto) AS total_deuda
        FROM alumnos
        JOIN pagos ON alumnos.id = pagos.alumno_id
        WHERE pagos.estado = 'Pendiente'
    """

    params = []

    if buscar:
        query += " AND (alumnos.nombres LIKE ? OR alumnos.apellidos LIKE ? OR alumnos.dni LIKE ?)"
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if grado:
        query += " AND alumnos.grado=?"
        params.append(grado)

    if seccion:
        query += " AND alumnos.seccion=?"
        params.append(seccion)

    if mes:
        query += " AND pagos.concepto=?"
        params.append(mes)

    query += """
        GROUP BY alumnos.id
        ORDER BY alumnos.grado, alumnos.seccion, alumnos.apellidos
    """

    cur.execute(query, params)
    deudores = cur.fetchall()

    conn.close()

    return render_template("deudores.html", deudores=deudores)

@app.route("/descargar_deudores_excel")
def descargar_deudores_excel():

    if "user" not in session:
        return redirect("/")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buscar = request.args.get("buscar", "")
    grado = request.args.get("grado", "")
    seccion = request.args.get("seccion", "")
    mes = request.args.get("mes", "")

    conn = get_db()
    cur = conn.cursor()

    conceptos = [
        "Matrícula", "Agenda", "Cuota de ingreso",
        "Marzo", "Abril", "Mayo", "Junio", "Julio",
        "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]

    query = """
        SELECT DISTINCT alumnos.id, alumnos.apellidos, alumnos.nombres, alumnos.dni, alumnos.grado, alumnos.seccion
        FROM alumnos
        JOIN pagos ON alumnos.id = pagos.alumno_id
        WHERE pagos.estado='Pendiente'
    """

    params = []

    if buscar:
        query += " AND (alumnos.nombres LIKE ? OR alumnos.apellidos LIKE ? OR alumnos.dni LIKE ?)"
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if grado:
        query += " AND alumnos.grado=?"
        params.append(grado)

    if seccion:
        query += " AND alumnos.seccion=?"
        params.append(seccion)

    if mes:
        query += " AND pagos.concepto=?"
        params.append(mes)

    query += " ORDER BY alumnos.grado, alumnos.seccion, alumnos.apellidos"

    cur.execute(query, params)
    alumnos = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Deudores"

    encabezados = ["DNI", "APELLIDOS", "NOMBRES", "GRADO", "SECCIÓN"] + conceptos
    ws.append(encabezados)

    for alumno in alumnos:
        alumno_id = alumno[0]

        fila = [
            alumno[3],
            alumno[1],
            alumno[2],
            alumno[4],
            alumno[5]
        ]

        for concepto in conceptos:
            cur.execute("""
                SELECT monto, estado
                FROM pagos
                WHERE alumno_id=? AND concepto=?
            """, (alumno_id, concepto))

            pago = cur.fetchone()

            if pago:
                monto = pago[0]
                estado = pago[1]

                if estado == "Pagado":
                    fila.append(f"PAGADO S/{monto:.2f}")
                else:
                    fila.append(f"DEBIENDO S/{monto:.2f}")
            else:
                fila.append("-")

        ws.append(fila)

    conn.close()

    azul = PatternFill("solid", fgColor="1E3A8A")
    verde = PatternFill("solid", fgColor="C6EFCE")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    blanco = Font(color="FFFFFF", bold=True)

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = azul
        cell.font = blanco
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if isinstance(cell.value, str):
                if cell.value.startswith("PAGADO"):
                    cell.fill = verde
                elif cell.value.startswith("DEBIENDO"):
                    cell.fill = rojo

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    nombre_archivo = "deudores.xlsx"
    wb.save(nombre_archivo)

    return send_file(nombre_archivo, as_attachment=True)

@app.route("/editar_alumno/<int:alumno_id>", methods=["GET", "POST"])
def editar_alumno(alumno_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    if request.method == "POST":
        nombres = request.form["nombres"]
        apellidos = request.form["apellidos"]
        dni = request.form["dni"]
        grado = request.form["grado"]
        seccion = request.form["seccion"]
        apoderado = request.form["apoderado"]
        telefono = request.form["telefono"]
        estado = request.form["estado"]

        cur.execute("""
            UPDATE alumnos
            SET nombres=?, apellidos=?, dni=?, grado=?, seccion=?, apoderado=?, telefono=?, estado=?
            WHERE id=?
        """, (nombres, apellidos, dni, grado, seccion, apoderado, telefono, estado, alumno_id))

        conn.commit()
        conn.close()
        return redirect("/alumnos")

    cur.execute("SELECT * FROM alumnos WHERE id=?", (alumno_id,))
    alumno = cur.fetchone()

    conn.close()

    return render_template("editar_alumno.html", alumno=alumno)


@app.route("/importar_txt")
def importar_txt():
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    with open("alumnos.txt", "r", encoding="utf-8") as f:
        lineas = f.readlines()

    grado_actual = ""
    seccion_actual = "A"

    for linea in lineas:
        linea = linea.strip()

        if not linea:
            continue

        linea_mayus = linea.upper()

        if linea_mayus.startswith("PRIMERO"):
            grado_actual = "1"
            seccion_actual = "A"
            continue
        elif linea_mayus.startswith("SEGUNDO"):
            grado_actual = "2"
            seccion_actual = "A"
            continue
        elif linea_mayus.startswith("TERCERO"):
            grado_actual = "3"
        elif linea_mayus.startswith("CUARTO"):
            grado_actual = "4"
        elif linea_mayus.startswith("QUINTO"):
            grado_actual = "5"
        else:
            partes = linea.split()

            if len(partes) < 3:
                continue

            apellidos = " ".join(partes[:2])
            nombres = " ".join(partes[2:])

            cur.execute("""
                SELECT id FROM alumnos
                WHERE nombres=? AND apellidos=? AND grado=? AND seccion=?
            """, (nombres, apellidos, grado_actual, seccion_actual))

            existe = cur.fetchone()

            if not existe:
                cur.execute("""
                    INSERT INTO alumnos (nombres, apellidos, grado, seccion)
                    VALUES (?, ?, ?, ?)
                """, (nombres, apellidos, grado_actual, seccion_actual))

            continue

        if '"A"' in linea_mayus:
            seccion_actual = "A"
        elif '"B"' in linea_mayus:
            seccion_actual = "B"
        elif '"C"' in linea_mayus:
            seccion_actual = "C"
        else:
            seccion_actual = "A"

    conn.commit()
    conn.close()

    return "Alumnos importados correctamente con grado y sección"


@app.route("/limpiar_alumnos")
def limpiar_alumnos():
    if "user" not in session:
        return redirect("/")

    if session.get("rol") != "admin":
        return "No tienes permiso para realizar esta acción", 403

    conn = get_db()
    cur = conn.cursor()

    cur.execute("DELETE FROM pagos")
    cur.execute("DELETE FROM matriculas")
    cur.execute("DELETE FROM alumnos")

    cur.execute("DELETE FROM sqlite_sequence WHERE name='pagos'")
    cur.execute("DELETE FROM sqlite_sequence WHERE name='matriculas'")
    cur.execute("DELETE FROM sqlite_sequence WHERE name='alumnos'")

    conn.commit()
    conn.close()

    return "Base de datos limpiada y IDs reiniciados desde 1"


@app.route("/alumno/<int:alumno_id>")
def detalle_alumno(alumno_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM alumnos WHERE id=?", (alumno_id,))
    alumno = cur.fetchone()

    cur.execute("""
        SELECT 
            pagos.concepto,
            pagos.monto,
            pagos.estado,
            pagos.fecha_pago,
            (
                SELECT boleta_detalles.boleta_id
                FROM boleta_detalles
                WHERE boleta_detalles.pago_id = pagos.id
                ORDER BY boleta_detalles.id DESC
                LIMIT 1
            ) AS boleta_id
        FROM pagos
        WHERE pagos.alumno_id=?
        ORDER BY pagos.id ASC
    """, (alumno_id,))
    pagos = cur.fetchall()

    cur.execute("SELECT SUM(monto) FROM pagos WHERE alumno_id=?", (alumno_id,))
    pendiente = cur.fetchone()[0] or 0

    try:
        cur.execute("SELECT SUM(monto_recibido) FROM boletas WHERE alumno_id=?", (alumno_id,))
        pagado = cur.fetchone()[0] or 0
    except:
        pagado = 0

    total = pendiente + pagado

    conn.close()

    return render_template(
        "detalle_alumno.html",
        alumno=alumno,
        pagos=pagos,
        total=total,
        pagado=pagado,
        pendiente=pendiente
    )


@app.route("/eliminar_alumno/<int:alumno_id>")
def eliminar_alumno(alumno_id):
    if "user" not in session:
        return redirect("/")

    if session.get("rol") != "admin":
        return "No tienes permiso para realizar esta acción", 403

    conn = get_db()
    cur = conn.cursor()

    cur.execute("DELETE FROM pagos WHERE alumno_id=?", (alumno_id,))
    cur.execute("DELETE FROM matriculas WHERE alumno_id=?", (alumno_id,))
    cur.execute("DELETE FROM alumnos WHERE id=?", (alumno_id,))

    conn.commit()
    conn.close()

    return redirect("/alumnos")

@app.route("/borrar_historial_matricula/<int:alumno_id>")
def borrar_historial_matricula(alumno_id):
    if "user" not in session:
        return redirect("/")

    if session.get("rol") != "admin":
        return "No tienes permiso para realizar esta acción", 403

    conn = get_db()
    cur = conn.cursor()

    cur.execute("DELETE FROM boleta_detalles WHERE pago_id IN (SELECT id FROM pagos WHERE alumno_id=?)", (alumno_id,))
    cur.execute("DELETE FROM boletas WHERE alumno_id=?", (alumno_id,))
    cur.execute("DELETE FROM pagos WHERE alumno_id=?", (alumno_id,))
    cur.execute("DELETE FROM matriculas WHERE alumno_id=?", (alumno_id,))

    conn.commit()
    conn.close()

    return redirect("/matricula")

@app.route("/anular_boleta/<int:boleta_id>", methods=["POST"])
def anular_boleta(boleta_id):
    if "user" not in session:
        return redirect("/")

    if session.get("rol") != "admin":
        return "No tienes permiso para realizar esta acción", 403

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT alumno_id FROM boletas WHERE id=?", (boleta_id,))
    boleta = cur.fetchone()

    if not boleta:
        conn.close()
        return redirect("/boletas")

    alumno_id = boleta[0]

    cur.execute("""
        SELECT pago_id, monto_aplicado
        FROM boleta_detalles
        WHERE boleta_id=?
    """, (boleta_id,))

    detalles = cur.fetchall()

    for pago_id, monto_aplicado in detalles:
        cur.execute("""
            UPDATE pagos
            SET monto = monto + ?,
                estado = 'Pendiente',
                fecha_pago = NULL
            WHERE id=?
        """, (monto_aplicado, pago_id))

    cur.execute("DELETE FROM boleta_detalles WHERE boleta_id=?", (boleta_id,))
    cur.execute("DELETE FROM boletas WHERE id=?", (boleta_id,))

    conn.commit()
    conn.close()

    return redirect("/boletas_alumno/" + str(alumno_id))


@app.route("/exonerar_pago/<int:pago_id>", methods=["POST"])
def exonerar_pago(pago_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT alumno_id
        FROM pagos
        WHERE id=?
    """, (pago_id,))
    pago = cur.fetchone()

    if not pago:
        conn.close()
        return redirect("/matricula")

    alumno_id = pago[0]
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cur.execute("""
        UPDATE pagos
        SET monto=0,
            estado='Exonerado',
            fecha_pago=?
        WHERE id=?
    """, (fecha, pago_id))

    conn.commit()
    conn.close()

    return redirect("/pagos/" + str(alumno_id))


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

@app.route("/reset_boletas")
def reset_boletas():

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("DELETE FROM boleta_detalles")
    except:
        pass

    try:
        cur.execute("DELETE FROM boletas")
    except:
        pass

    try:
        cur.execute("""
            UPDATE pagos
            SET estado='Pendiente',
                fecha_pago=NULL
        """)
    except:
        pass

    conn.commit()
    conn.close()

    return "Boletas reiniciadas correctamente 😎"

if __name__ == "__main__":
    app.run(debug=True)